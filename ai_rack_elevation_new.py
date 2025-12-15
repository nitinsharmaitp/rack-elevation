import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
import base64

# Set default datacenter background image
def set_default_bg():
    try:
        with open("datacenter_default.jpg", "rb") as img_file:
            b64 = base64.b64encode(img_file.read()).decode()
        style = f"""
            <style>
            .stApp {{
                background-image: url("data:image/jpeg;base64,{b64}");
                background-position: center center;
                background-repeat: no-repeat;
                background-size: contain;      /* or 'auto'/'1200px auto' as you wish */
                background-attachment: fixed;
            }}
            </style>
        """
        st.markdown(style, unsafe_allow_html=True)
    except Exception:
        pass  # Ignore errors if image file missing

st.set_page_config(layout="wide")
set_default_bg()

st.title("VizRack")
st.caption("Your Infrastructure. Visually Elevated")

with st.expander("📝 **What is VizRack?**", expanded=True):
    st.markdown("""
    **Quick Features:**
    - Automated, color-coded rack elevation diagrams from Excel input
    - Per-cabinet rack height and accurate device placement
    - Multi-device & multi-RU handling, instantly visual
    - Role-based device coloring, empty RU marking
    - Full version history and input inventory preserved
    - Download ready-to-share Excel file with all diagrams and legends

    **Output:**
    - Multi-sheet Excel (.xlsx) file with Version History, Inventory, Rack Elevations, Color Legend
    - Professionally formatted, ready for reviews and audits
    """)

DEVICE_COLORS = [
    ("Switch", "ADD8E6"),  # soft blue
    ("Firewall", "FFFF00"),
    ("Router", "00B0F0"),
    ("Server", "92D050"),
    ("Storage", "7030A0"),
    ("UPS", "C00000"),
    ("PDU", "7F7F7F"),
    ("KVM", "00B050"),
    ("Patch Panel", "A6A6A6"),
    ("Other", "D9D9D9"),
]

DEFAULT_FILL = "FFFFFF"
SHEET_TAB_COLOR = "B7E1CD"  # light green tab

def get_color_fill(role):
    if not isinstance(role, str):
        return PatternFill("solid", fgColor=DEFAULT_FILL)
    for dev_type, color in DEVICE_COLORS:
        if dev_type.lower() == role.strip().lower():
            return PatternFill("solid", fgColor=color)
    return PatternFill("solid", fgColor=DEFAULT_FILL)

def parse_input(file):
    df = pd.read_excel(file, engine="openpyxl")
    required_columns = ["Rack ID", "RU Position", "Units", "Device Name", "IP Address",
                        "Device Role", "Model", "Serial Number", "Power Supply Count", "PL", "PR"]
    missing_cols = [c for c in required_columns if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing columns: {missing_cols}")
    if "RU Height" not in df.columns:
        df["RU Height"] = None
    return df

def parse_rack_id(rack_id):
    m = re.match(r"Row\s*([^\s]+)\s+Cabinet\s*([^\s]+)", str(rack_id), re.I)
    if m:
        return f"Row{m.group(1)}", f"Cabinet{m.group(2)}"
    return "RowUnknown", "CabinetUnknown"

def parse_rus(ru_val):
    if pd.isna(ru_val):
        return []
    s = str(ru_val).replace(" ", "")
    rus = []
    for part in s.split(","):
        if "-" in part:
            try:
                start, end = map(int, part.split("-"))
                rus.extend(range(min(start, end), max(start, end) + 1))
            except:
                continue
        elif part.isdigit():
            rus.append(int(part))
    return sorted(set(rus), reverse=True)

def clean_text(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().upper() == "NAN":
        return "NA"
    return str(val).strip()

def clean_power(val):
    if pd.isna(val) or str(val).strip() == "":
        return "N"
    return str(val).strip().upper()

def safe_int(val, default=1):
    try:
        return int(float(val))
    except:
        return default

def auto_adjust_cols(ws, extra=3):
    for i, col in enumerate(ws.columns, 1):
        max_len = 1
        col_letter = get_column_letter(i)
        for cell in col:
            try:
                if cell.value:
                    lines = str(cell.value).splitlines()
                    max_len = max(max_len, max(len(line) for line in lines))
            except:
                continue
        ws.column_dimensions[col_letter].width = max_len + extra

def row_sort_key(row_key):
    m = re.search(r'Row(\d+)', str(row_key))
    if m:
        try:
            return int(m.group(1))
        except:
            return float('inf')
    return float('inf')

def cabinet_sort_key(cab_key):
    m = re.search(r'Cabinet\s*(\d+)', str(cab_key), re.I)
    if m:
        try:
            return int(m.group(1))
        except:
            return float('inf')
    return float('inf')

def build_ru_map(df):
    ru_map = {}
    for _, row in df.iterrows():
        key = parse_rack_id(row["Rack ID"])
        if pd.notna(row.get("RU Height")) and str(row["RU Height"]).strip():
            val = safe_int(row["RU Height"])
            if val:
                ru_map[key] = val
    return ru_map

def create_rack_layout(devices, height):
    rack = [None] * height
    skip_set = set()
    for d in devices:
        rus = parse_rus(d.get("RU Position"))
        if not rus:
            continue
        span = len(rus)
        top = max(rus)
        pos = height - top
        if pos in skip_set:
            continue
        rack[pos] = d
        for i in range(1, span):
            if (pos + i) < height:
                rack[pos + i] = "_skip_"
                skip_set.add(pos + i)
    return rack

def add_version_history(wb, engineer):
    ws = wb.create_sheet("Version History", 0)
    headers = ["Version", "Date", "Updated By", "Reason for Change"]
    data = [[1, datetime.now().strftime("%d-%m-%Y"), engineer, "Initial Draft"]]

    ws.merge_cells("A1:D1")
    cell = ws.cell(1, 1, "Version Control:")
    cell.font = Font(size=14, bold=True, color="006400")
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.fill = PatternFill('solid', fgColor='B7E1CD')
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill('solid', fgColor='00B050')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    for i, h in enumerate(headers, 1):
        c = ws.cell(2, i, h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 25

    for r, row in enumerate(data, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = Font(bold = (c == 1))
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if c == 1:
                cell.number_format = 'General'

    for r in range(3 + len(data), 10):
        for c in range(1, len(headers)+1):
            cell = ws.cell(r, c, "")
            cell.border = border

    auto_adjust_cols(ws, extra=8)
    ws.sheet_properties.tabColor = SHEET_TAB_COLOR

def add_inventory_sheet(wb, df, fileobj):
    ws = wb.create_sheet("Rack Inventory", 1)
    try:
        src_wb = load_workbook(fileobj)
        src_ws = src_wb[src_wb.sheetnames[0]]
        for i, col in enumerate(df.columns, 1):
            cell = ws.cell(1, i, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
            src_cell = src_ws.cell(1, i)
            if src_cell.fill and src_cell.fill.fgColor.type == 'rgb' and src_cell.fill.fgColor.rgb != '00000000':
                cell.fill = PatternFill('solid', fgColor=src_cell.fill.fgColor.rgb)
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(r, c, val if pd.notna(val) else "")
                src_cell = src_ws.cell(r, c)
                cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
                if src_cell.fill and src_cell.fill.fgColor.type == 'rgb' and src_cell.fill.fgColor.rgb != '00000000':
                    cell.fill = PatternFill('solid', fgColor=src_cell.fill.fgColor.rgb)
                if hasattr(src_cell, "font"):
                    cell.font = copy(src_cell.font)
    except Exception:
        # Fallback plain copy
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        for i, col in enumerate(df.columns, 1):
            cell = ws.cell(1, i, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, val in enumerate(row, 1):
                ws.cell(r, c, val if pd.notna(val) else "").border = border
    auto_adjust_cols(ws, extra=5)
    ws.sheet_properties.tabColor = SHEET_TAB_COLOR

def add_color_legend(wb):
    ws = wb.create_sheet("Device Color Legend")
    headers = ["Device Type", "Color Code", "Preview"]
    header_fill = PatternFill('solid', fgColor='BDD7EE')
    header_font = Font(bold=True)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r, (dev, col) in enumerate(DEVICE_COLORS, 2):
        c1 = ws.cell(r, 1, dev)
        c1.border = border
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c2 = ws.cell(r, 2, f"#{col}")
        c2.border = border
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c3 = ws.cell(r, 3)
        c3.fill = PatternFill('solid', fgColor=col)
        c3.border = border

    auto_adjust_cols(ws, extra=5)
    ws.sheet_properties.tabColor = SHEET_TAB_COLOR

def create_rack_excel(df, input_df, engineer_name, input_file):
    structure = {}
    for _, row in df.iterrows():
        row_k, cab_k = parse_rack_id(row["Rack ID"])
        if row_k == "RowUnknown":
            continue
        structure.setdefault(row_k, {})
        structure[row_k].setdefault(cab_k, [])
        structure[row_k][cab_k].append(row.to_dict())

    ru_map = build_ru_map(df)
    default_ru_height = 42

    wb = Workbook()
    wb.remove(wb.active)

    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    dev_font = Font(name="Calibri", size=9, bold=True)

    add_version_history(wb, engineer_name)
    add_inventory_sheet(wb, input_df, input_file)

    rows_sorted = sorted(structure.keys(), key=row_sort_key)
    sheet_idx = 2

    for row_key in rows_sorted:
        ws = wb.create_sheet(row_key, sheet_idx)
        sheet_idx += 1
        ws.sheet_properties.tabColor = SHEET_TAB_COLOR

        cabinets_sorted = sorted(structure[row_key].items(), key=lambda x: cabinet_sort_key(x[0]))

        for i, (cab_name, devices) in enumerate(cabinets_sorted):
            col_start = 1 + i*5
            ru_height = ru_map.get((row_key, cab_name), default_ru_height)

            # Header styling
            ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start+3)
            h_cell = ws.cell(1, col_start)
            h_cell.value = cab_name
            h_cell.font = Font(bold=True, size=12, color="18453B")
            h_cell.fill = PatternFill("solid", fgColor="BDD7EE")
            h_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 28

            # Sub-headers
            subheaders = ["RU", "PL", "Device", "PR"]
            for offset, sh in enumerate(subheaders):
                c = ws.cell(2, col_start + offset)
                c.value = sh
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="BDD7EE")
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[2].height = 24

            # Row heights for data rows
            for r in range(3, 3 + ru_height):
                ws.row_dimensions[r].height = 40

            # Create rack layout
            layout = create_rack_layout(devices, ru_height)
            row_ptr = 0

            while row_ptr < ru_height:
                cur_row = 3 + row_ptr
                slot = layout[row_ptr]

                if slot == "_skip_":
                    row_ptr += 1
                    continue

                if slot:
                    unit_cnt = safe_int(slot.get("Units", 1))
                    if unit_cnt < 1:
                        unit_cnt = 1

                    end_row = cur_row + unit_cnt - 1

                    # Merge RU cell and set value
                    ws.merge_cells(start_row=cur_row, start_column=col_start, end_row=end_row, end_column=col_start)
                    ru_cell = ws.cell(cur_row, col_start)
                    rus = parse_rus(slot.get("RU Position"))
                    ru_label = f"{max(rus)}-{min(rus)}" if len(rus) > 1 else str(rus[0]) if rus else slot.get("RU Position", "")
                    ru_cell.value = ru_label
                    ru_cell.font = Font(bold=True)
                    ru_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ru_cell.border = border

                    # Merge PL, Device, PR cells and set values
                    for c_offset, key in zip([1,2,3], ["PL", "Device", "PR"]):
                        ws.merge_cells(start_row=cur_row, start_column=col_start + c_offset, end_row=end_row, end_column=col_start + c_offset)
                        cell = ws.cell(cur_row, col_start + c_offset)
                        val = slot.get(key, "")
                        if key == "Device":
                            ip = slot.get("IP Address", "")
                            sn = slot.get("Serial Number", "")
                            val = f"{clean_text(val)}\n{clean_text(ip)}\nSN: {clean_text(sn)}"
                            cell.font = dev_font
                            cell.fill = get_color_fill(slot.get("Device Role", ""))
                        else:
                            val = clean_power(val)
                            green_fill = PatternFill("solid", fgColor="C6EFCE")
                            red_fill = PatternFill("solid", fgColor="FFC7CE")
                            cell.fill = green_fill if (slot.get("PL", "N") == "Y" and slot.get("PR", "N") == "Y") else red_fill
                            cell.font = Font(bold=True)
                        cell.value = val
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cell.border = border

                    # Add border to entire merged region
                    for r in range(cur_row, end_row + 1):
                        for c in range(col_start, col_start + 4):
                            ws.cell(r, c).border = border

                    row_ptr += unit_cnt
                else:
                    ru_cell = ws.cell(cur_row, col_start)
                    ru_cell.value = ru_height - row_ptr
                    ru_cell.font = Font(bold=True)
                    ru_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ru_cell.border = border

                    ws.cell(cur_row, col_start + 1, "").border = border
                    empty_dev_cell = ws.cell(cur_row, col_start + 2, "Empty/No Device")
                    empty_dev_cell.font = Font(italic=True, color="888888")
                    empty_dev_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    empty_dev_cell.border = border
                    ws.cell(cur_row, col_start + 3, "").border = border

                    row_ptr += 1

            auto_adjust_cols(ws, extra=9)
    
    add_color_legend(wb)

    for sheet in wb.worksheets:
        if not sheet.sheet_properties.tabColor:
            sheet.sheet_properties.tabColor = SHEET_TAB_COLOR

    if "Sheet" in wb.sheetnames:  # Remove default sheet if present
        wb.remove(wb["Sheet"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Streamlit UI
with st.form("form"):
    excel_file = st.file_uploader("Upload Excel File for Rack Data (.xlsx)", type=['xlsx'])
    engineer_name = st.text_input("Engineer Name", value="", placeholder="Enter your name")
    submitted = st.form_submit_button("Generate VizRack")

if submitted:
    if not excel_file:
        st.error("Please upload a valid Excel file.")
    else:
        try:
            df = parse_input(excel_file)
            output = create_rack_excel(df, df, engineer_name, excel_file)
            st.success("Rack elevation diagram generated!")
            st.download_button("Download VizRack Output Excel",
                               data=output,
                               file_name="VizRack_Rack_Elevation.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Error: {e}")
